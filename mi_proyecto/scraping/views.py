from django.shortcuts import render, get_object_or_404, redirect
from .models import Entry, PersonalInfo, Investigacion, FormacionAcademica, ExperienciaProfesional
from .forms import EntryForm
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from django.core.management import call_command
from openpyxl import Workbook
import datetime
import subprocess
import os
import json

# 📌 Página de inicio
def home_page(request):
    return render(request, 'home.html')

# 📌 Base
def base(request):
    return render(request, 'base.html', {'entry': "Some entry data"})

# 📌 Vista de Ingeniería de Sistemas
def ingenieria_sistemas_view(request):
    json_file_path = os.path.join(os.path.dirname(__file__), '..', 'data.json')
    try:
        with open(json_file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
    except FileNotFoundError:
        data = []
    return render(request, 'ingenieria_sistemas.html', {'data': data})

# 📌 Lista de docentes con búsqueda
def entry_list(request):
    query = request.GET.get("q")
    entries = Entry.objects.filter(name__icontains=query) if query else Entry.objects.all()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('scraping/_entry_list_items.html', {'entries': entries})
        return JsonResponse({'html': html})

    return render(request, 'scraping/entry_list.html', {'entries': entries})

# 📌 Vista detallada de un docente
def entry_detail(request, pk):
    entry = get_object_or_404(Entry, pk=pk)
    personal_info = PersonalInfo.objects.filter(entry=entry).first()
    formacion = FormacionAcademica.objects.filter(entry=entry).first()
    experiencia = ExperienciaProfesional.objects.filter(entry=entry).first()
    investigaciones = Investigacion.objects.filter(entry=entry).first()

    return render(request, 'scraping/entry_detail.html', {
        'entry': entry,
        'personal_info': personal_info,
        'formacion': formacion,
        'experiencia': experiencia,
        'investigaciones': investigaciones,
    })

# 📌 Crear entrada
def entry_new(request):
    if request.method == "POST":
        form = EntryForm(request.POST)
        if form.is_valid():
            entry = form.save()
            return redirect('entry_detail', pk=entry.pk)
    else:
        form = EntryForm()
    return render(request, 'scraping/entry_edit.html', {'form': form})

# 📌 Editar entrada
def entry_edit(request, pk):
    entry = get_object_or_404(Entry, pk=pk)
    if request.method == "POST":
        form = EntryForm(request.POST, instance=entry)
        if form.is_valid():
            entry = form.save()
            return redirect('entry_detail', pk=entry.pk)
    else:
        form = EntryForm(instance=entry)
    return render(request, 'scraping/entry_edit.html', {'form': form})

# 📌 Eliminar entrada
def entry_delete(request, pk):
    entry = get_object_or_404(Entry, pk=pk)
    entry.delete()
    return redirect('entry_list')

# 📌 Exportar JSON de todos los docentes
def extracted_data_list(request):
    entries = Entry.objects.all()
    data = []

    for entry in entries:
        personal_info = PersonalInfo.objects.filter(entry=entry).first()
        investigacion = Investigacion.objects.filter(entry=entry).first()

        data.append({
            'name': entry.name,
            'href': entry.href,
            'personal_info': {
                'nombre': personal_info.nombre if personal_info else None,
                'nombre_citaciones': personal_info.nombre_citaciones if personal_info else None,
                'categoria': personal_info.categoria if personal_info else None,
                'par_evaluador': personal_info.par_evaluador if personal_info else None,
                'nacionalidad': personal_info.nacionalidad if personal_info else None,
                'sexo': personal_info.sexo if personal_info else None,
            } if personal_info else None,
            'investigaciones': investigacion.datos if investigacion and investigacion.datos else {}
        })

    return JsonResponse(data, safe=False)

# 📌 Exportar a Excel dinámico
@csrf_exempt
def exportar_datos_view(request):
    if request.method == 'POST':
        incluir_personal = 'incluir_personal' in request.POST
        incluir_formacion = 'incluir_formacion' in request.POST
        incluir_experiencia = 'incluir_experiencia' in request.POST
        incluir_investigaciones = 'incluir_investigaciones' in request.POST
        estructura = request.POST.get('estructura', 'una_hoja')  # Default: una sola hoja

        wb = Workbook()
        wb.remove(wb.active)

        entries = Entry.objects.all()

        if estructura == 'una_hoja':
            ws = wb.create_sheet("Docentes")
            encabezados = ['Nombre', 'URL']
            if incluir_personal:
                encabezados += ['Nombre en Citaciones', 'Categoría', 'Par Evaluador', 'Nacionalidad', 'Sexo']
            if incluir_formacion:
                encabezados += ['Nivel Académico', 'Institución']
            if incluir_experiencia:
                encabezados += ['Institución (Trabajo)', 'Cargo', 'Desde', 'Hasta']
            if incluir_investigaciones:
                encabezados += ['Investigaciones (Resumen)']
            ws.append(encabezados)

            for entry in entries:
                fila = [entry.name, entry.href]
                if incluir_personal:
                    info = PersonalInfo.objects.filter(entry=entry).first()
                    fila += [
                        info.nombre_citaciones if info else '',
                        info.categoria if info else '',
                        'Sí' if info and info.par_evaluador else 'No',
                        info.nacionalidad if info else '',
                        info.sexo if info else '',
                    ]
                if incluir_formacion:
                    formacion = FormacionAcademica.objects.filter(entry=entry).first()
                    fila += [
                        formacion.nivel if formacion else '',
                        formacion.institucion if formacion else '',
                    ]
                if incluir_experiencia:
                    exp = ExperienciaProfesional.objects.filter(entry=entry).first()
                    fila += [
                        exp.institucion if exp else '',
                        exp.cargo if exp else '',
                        exp.desde if exp else '',
                        exp.hasta if exp else '',
                    ]
                if incluir_investigaciones:
                    inv = Investigacion.objects.filter(entry=entry).first()
                    resumen = ', '.join([f"{k}: {v}" for k, v in inv.datos.items()]) if inv and inv.datos else ''
                    fila.append(resumen)
                ws.append(fila)

        elif estructura == 'una_por_docente':
            for entry in entries:
                ws = wb.create_sheet(entry.name[:31])
                ws.append(['Campo', 'Valor'])

                ws.append(['Nombre', entry.name])
                ws.append(['URL', entry.href])

                if incluir_personal:
                    info = PersonalInfo.objects.filter(entry=entry).first()
                    ws.append(['Nombre en Citaciones', info.nombre_citaciones if info else ''])
                    ws.append(['Categoría', info.categoria if info else ''])
                    ws.append(['Par Evaluador', 'Sí' if info and info.par_evaluador else 'No'])
                    ws.append(['Nacionalidad', info.nacionalidad if info else ''])
                    ws.append(['Sexo', info.sexo if info else ''])

                if incluir_formacion:
                    formacion = FormacionAcademica.objects.filter(entry=entry).first()
                    ws.append(['Nivel Académico', formacion.nivel if formacion else ''])
                    ws.append(['Institución', formacion.institucion if formacion else ''])

                if incluir_experiencia:
                    exp = ExperienciaProfesional.objects.filter(entry=entry).first()
                    ws.append(['Institución (Trabajo)', exp.institucion if exp else ''])
                    ws.append(['Cargo', exp.cargo if exp else ''])
                    ws.append(['Desde', exp.desde if exp else ''])
                    ws.append(['Hasta', exp.hasta if exp else ''])

                if incluir_investigaciones:
                    inv = Investigacion.objects.filter(entry=entry).first()
                    if inv and inv.datos:
                        for k, v in inv.datos.items():
                            ws.append([f"{k}", v])

        elif estructura == 'una_por_tipo':
            if incluir_personal:
                ws = wb.create_sheet("Información Personal")
                ws.append(['Nombre', 'Nombre en Citaciones', 'Categoría', 'Par Evaluador', 'Nacionalidad', 'Sexo'])
                for entry in entries:
                    info = PersonalInfo.objects.filter(entry=entry).first()
                    ws.append([
                        entry.name,
                        info.nombre_citaciones if info else '',
                        info.categoria if info else '',
                        'Sí' if info and info.par_evaluador else 'No',
                        info.nacionalidad if info else '',
                        info.sexo if info else '',
                    ])

            if incluir_formacion:
                ws = wb.create_sheet("Formación Académica")
                ws.append(['Nombre', 'Nivel Académico', 'Institución'])
                for entry in entries:
                    formacion = FormacionAcademica.objects.filter(entry=entry).first()
                    ws.append([
                        entry.name,
                        formacion.nivel if formacion else '',
                        formacion.institucion if formacion else '',
                    ])

            if incluir_experiencia:
                ws = wb.create_sheet("Experiencia Profesional")
                ws.append(['Nombre', 'Institución', 'Cargo', 'Desde', 'Hasta'])
                for entry in entries:
                    exp = ExperienciaProfesional.objects.filter(entry=entry).first()
                    ws.append([
                        entry.name,
                        exp.institucion if exp else '',
                        exp.cargo if exp else '',
                        exp.desde if exp else '',
                        exp.hasta if exp else '',
                    ])

            if incluir_investigaciones:
                ws = wb.create_sheet("Investigaciones")
                ws.append(['Nombre', 'Resumen'])
                for entry in entries:
                    inv = Investigacion.objects.filter(entry=entry).first()
                    resumen = ', '.join([f"{k}: {v}" for k, v in inv.datos.items()]) if inv and inv.datos else ''
                    ws.append([entry.name, resumen])

        # Preparar para descarga
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        nombre_archivo = f"docentes_exportados_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        wb.save(response)
        return response

    return render(request, 'scraping/exportar_datos.html')

# 📌 Ejecutar scraping desde la vista
def ejecutar_scraping(request):
    try:
        call_command('extract_data')
        return JsonResponse({'mensaje': 'Scraping ejecutado exitosamente.'})
    except Exception as e:
        return JsonResponse({'mensaje': f'Error durante el scraping: {str(e)}'}, status=500)

# 📌 Transmitir consola de scraping en tiempo real (SSE si se activa)
def iniciar_scraping_view(request):
    def generate():
        process = subprocess.Popen(
            ["python3", "manage.py", "extract_data"],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        for line in iter(process.stdout.readline, ''):
            yield f"data:{line.strip()}\n\n"
        process.stdout.close()

    return StreamingHttpResponse(generate(), content_type='text/event-stream')
